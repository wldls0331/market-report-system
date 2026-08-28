from __future__ import annotations

import datetime as dt
import re
import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from config.cell_mapping import (
    COMMENT_FIELDS,
    DATE_CELLS,
    EXPECTED_PDF_PAGES,
    FORMULA_CELLS,
    KOREA_COMMENT_SOURCE_KEY,
    KOREA_WORLDWIDE_COMMENT_KEY,
    NUMBER_FIELDS,
    REQUIRED_LABELS,
    SALES_ASSUMPTION_HEADER_CELL,
    SALES_ASSUMPTION_HEADER_TEMPLATE,
    SHEET_NAME_PATTERN,
    STRATEGY_FIELDS,
    TITLE_CELL,
    WORLDWIDE_NEW_POINT_CELLS,
    all_mapped_cells,
    chart_window_cells,
)
from config.paths import OUTPUT_DIR, find_template, session_output_dir
from services.chart_service import shift_weekly_chart_window
from services.comment_service import TBN, compose_comment_lines, compose_strategy_text
from services.pdf_service import (
    PdfExportError,
    excel_com_available,
    export_report_pdf_cloud,
    prepare_sheet_and_export_pdf,
)
from services.pricing_month_service import YearMonth, default_pricing_month
from services.working_day_service import (
    WorkingDayError,
    format_output_stem,
    format_pdf_filename,
    format_report_title,
    format_sheet_name,
    previous_week_last_working_day,
    validate_report_date,
)

SHEET_RE = re.compile(SHEET_NAME_PATTERN)

# Excel header text stays Korean; Streamlit error copy uses English names.
_LABEL_UI_NAME = {
    "페이퍼": "Paper / MOPS",
    "시장가": "Bunker Market Price",
}


def _expected_label_ui(expected: str) -> str:
    return _LABEL_UI_NAME.get(expected, expected)


class GenerateError(RuntimeError):
    pass


@dataclass
class SheetInfo:
    name: str
    date: dt.date
    index: int  # 0-based


@dataclass
class GenerateResult:
    excel_path: Path
    pdf_path: Path | None
    sheet_name: str
    source_path: Path
    previous_sheet: str
    warnings: list[str] = field(default_factory=list)
    pdf_page_count: int | None = None
    pricing_month: str | None = None
    is_update: bool = False


def parse_sheet_date(name: str) -> dt.date | None:
    match = SHEET_RE.match(name.strip())
    if not match:
        return None
    year, month, day = (int(part) for part in match.groups())
    year += 2000
    try:
        return dt.date(year, month, day)
    except ValueError:
        return None


def list_report_sheets(sheet_names: list[str]) -> list[SheetInfo]:
    sheets: list[SheetInfo] = []
    for index, name in enumerate(sheet_names):
        date = parse_sheet_date(name)
        if date is None:
            continue
        sheets.append(SheetInfo(name=name, date=date, index=index))
    sheets.sort(key=lambda item: item.date, reverse=True)
    return sheets


def find_previous_sheet(sheet_names: list[str], report_date: dt.date) -> SheetInfo | None:
    for sheet in list_report_sheets(sheet_names):
        if sheet.date < report_date:
            return sheet
    return None


def sheet_exists_for_date(sheet_names: list[str], report_date: dt.date) -> str | None:
    expected = format_sheet_name(report_date)
    for name in sheet_names:
        parsed = parse_sheet_date(name)
        if parsed == report_date or name.startswith(expected):
            return name
    return None


def output_excel_path(report_date: dt.date) -> Path:
    return session_output_dir() / f"{format_output_stem(report_date)}.xlsx"


def report_exists_for_date(report_date: dt.date) -> bool:
    """True when the dated output workbook (or source) already has that day's sheet."""
    excel_path = output_excel_path(report_date)
    if excel_path.exists():
        try:
            if sheet_exists_for_date(_sheet_names(excel_path), report_date):
                return True
        except Exception:
            pass
    try:
        source = resolve_source_workbook()
        return sheet_exists_for_date(_sheet_names(source), report_date) is not None
    except Exception:
        return False


def resolve_source_workbook() -> Path:
    template = find_template()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    outputs = [
        path
        for path in OUTPUT_DIR.glob("*_Market Report.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not outputs:
        return template
    latest_output = max(outputs, key=lambda path: path.stat().st_mtime)
    try:
        template_sheets = _sheet_names(template)
        output_sheets = _sheet_names(latest_output)
        template_latest = list_report_sheets(template_sheets)
        output_latest = list_report_sheets(output_sheets)
        if output_latest and template_latest and output_latest[0].date >= template_latest[0].date:
            return latest_output
    except Exception:
        return template
    return template


def _sheet_names(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _cell_value(worksheet, address: str) -> Any:
    return worksheet[address].value


def load_previous_values(report_date: dt.date) -> dict[str, Any]:
    source = resolve_source_workbook()
    workbook = load_workbook(source, data_only=True)
    try:
        previous = find_previous_sheet(list(workbook.sheetnames), report_date)
        if previous is None:
            return {}
        worksheet = workbook[previous.name]
        values: dict[str, Any] = {"_previous_sheet": previous.name, "_source": str(source)}
        for item in NUMBER_FIELDS:
            values[item.key] = _cell_value(worksheet, item.cell)
        for item in COMMENT_FIELDS:
            parts = []
            for cell in item.cells:
                raw = _cell_value(worksheet, cell)
                if raw is None:
                    continue
                text = str(raw).strip()
                if text:
                    parts.append(text)
            values[item.key] = "\n".join(parts)
        for item in STRATEGY_FIELDS:
            raw = _cell_value(worksheet, item.cell)
            values[item.key] = "" if raw is None else str(raw)
        return values
    finally:
        workbook.close()


def parse_optional_number(raw: str | None) -> float | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text or text.upper() == TBN:
        return None
    text = text.replace(",", "")
    try:
        return float(text)
    except ValueError as exc:
        raise GenerateError(f"Invalid number: {raw}") from exc


def _format_paper_label(value: float | None) -> str:
    if value is None:
        return TBN
    if value == int(value):
        return f"{int(value):.2f}"
    return f"{value:.2f}"


def _validate_mapping(worksheet) -> None:
    missing = []
    for address in all_mapped_cells():
        try:
            worksheet[address]
        except Exception:
            missing.append(address)
    if missing:
        raise GenerateError(f"Required mapping cell is missing: {', '.join(missing)}")

    mismatches = []
    for address, expected in REQUIRED_LABELS.items():
        actual = worksheet[address].value
        if actual is None:
            mismatches.append(f"{address} (empty, expected '{_expected_label_ui(expected)}')")
            continue
        if str(actual).strip() != expected:
            mismatches.append(f"{address}='{actual}' (expected '{_expected_label_ui(expected)}')")
    if mismatches:
        raise GenerateError("Sheet structure does not match mapping: " + "; ".join(mismatches))


def _write_number(worksheet, address: str, value: float | None) -> None:
    cell = worksheet[address]
    if value is None:
        cell.value = None
    else:
        cell.value = value


def _write_text(worksheet, address: str, value: str) -> None:
    worksheet[address].value = value if value else None


def _shift_charts_openpyxl(worksheet, data_reference_date: dt.date) -> None:
    shift_weekly_chart_window(
        lambda address: worksheet[address].value,
        lambda address, value: setattr(worksheet[address], "value", value),
        data_reference_date,
    )


def _write_worldwide_new_point(write_number, inputs: dict[str, Any]) -> None:
    for address, key in WORLDWIDE_NEW_POINT_CELLS.items():
        write_number(address, parse_optional_number(inputs.get(key)))


def _write_typed_openpyxl(worksheet, address: str, value: Any) -> None:
    if address in FORMULA_CELLS:
        return
    if value is None or value == "":
        worksheet[address].value = None
        return
    if isinstance(value, dt.date) and not isinstance(value, dt.datetime):
        worksheet[address].value = dt.datetime.combine(value, dt.time())
        return
    worksheet[address].value = value


def _write_typed_com(com_ws, address: str, value: Any) -> None:
    if address in FORMULA_CELLS:
        return
    if isinstance(value, dt.date) and not isinstance(value, dt.datetime):
        com_ws.Range(address).Value = _excel_serial_date(value)
        return
    if isinstance(value, dt.datetime):
        com_ws.Range(address).Value = _excel_serial_date(value.date())
        return
    _write_com_value(com_ws, address, value)


def _overlay_extra_cells_openpyxl(worksheet, extra_cells: dict[str, Any] | None) -> None:
    if not extra_cells:
        return
    for address, value in extra_cells.items():
        _write_typed_openpyxl(worksheet, address, value)


def _overlay_extra_cells_com(com_ws, extra_cells: dict[str, Any] | None) -> None:
    if not extra_cells:
        return
    for address, value in extra_cells.items():
        _write_typed_com(com_ws, address, value)


def _apply_values(
    worksheet,
    report_date: dt.date,
    inputs: dict[str, Any],
    pricing_month: YearMonth,
    data_reference_date: dt.date,
    extra_cells: dict[str, Any] | None = None,
) -> None:
    extra_cells = extra_cells or {}
    skip_shift = any(address in extra_cells for address in chart_window_cells())
    if not skip_shift:
        _shift_charts_openpyxl(worksheet, data_reference_date)

    reference_dt = dt.datetime.combine(data_reference_date, dt.time())
    worksheet[DATE_CELLS["paper_date"]].value = reference_dt
    worksheet[DATE_CELLS["paper_table_date"]].value = reference_dt
    worksheet[DATE_CELLS["market_table_date"]].value = reference_dt
    worksheet[DATE_CELLS["premium_table_date"]].value = reference_dt
    worksheet[TITLE_CELL].value = format_report_title(report_date)

    for item in NUMBER_FIELDS:
        value = parse_optional_number(inputs.get(item.key))
        _write_number(worksheet, item.cell, value)
        if item.label_cell and item.label_template:
            rendered = item.label_template.format(
                month=pricing_month.abbr,
                value=_format_paper_label(value),
            )
            _write_text(worksheet, item.label_cell, rendered)

    if not skip_shift:
        _write_worldwide_new_point(
            lambda address, value: _write_number(worksheet, address, value),
            inputs,
        )

    comment_inputs = dict(inputs)
    worldwide = str(inputs.get(KOREA_WORLDWIDE_COMMENT_KEY, "") or "").strip()
    comment_inputs[KOREA_WORLDWIDE_COMMENT_KEY] = worldwide or inputs.get(KOREA_COMMENT_SOURCE_KEY, "")
    for item in COMMENT_FIELDS:
        lines = compose_comment_lines(comment_inputs.get(item.key, ""), len(item.cells))
        for address, line in zip(item.cells, lines):
            _write_text(worksheet, address, line)

    for item in STRATEGY_FIELDS:
        _write_text(worksheet, item.cell, compose_strategy_text(item.prefix, inputs.get(item.key, "")))

    worksheet[SALES_ASSUMPTION_HEADER_CELL].value = SALES_ASSUMPTION_HEADER_TEMPLATE.format(
        month=pricing_month.full_name
    )
    _overlay_extra_cells_openpyxl(worksheet, extra_cells)


def _excel_serial_date(date: dt.date) -> int:
    return (dt.datetime(date.year, date.month, date.day) - dt.datetime(1899, 12, 30)).days


def _write_com_value(com_ws, address: str, value: Any) -> None:
    rng = com_ws.Range(address)
    if value is None or value == "":
        rng.ClearContents()
        return
    rng.Value = value


def _write_report_title_com(com_ws, report_date: dt.date) -> None:
    rng = com_ws.Range(TITLE_CELL)
    rng.Value = format_report_title(report_date)
    rng.WrapText = False
    rng.Font.Bold = True


def _apply_values_com(
    com_ws,
    report_date: dt.date,
    inputs: dict[str, Any],
    pricing_month: YearMonth,
    data_reference_date: dt.date,
    extra_cells: dict[str, Any] | None = None,
) -> None:
    extra_cells = extra_cells or {}
    skip_shift = any(address in extra_cells for address in chart_window_cells())

    def get_value(address: str):
        return com_ws.Range(address).Value

    def set_value(address: str, value: Any) -> None:
        _write_typed_com(com_ws, address, value)

    if not skip_shift:
        shift_weekly_chart_window(get_value, set_value, data_reference_date)

    reference_serial = _excel_serial_date(data_reference_date)
    com_ws.Range(DATE_CELLS["paper_date"]).Value = reference_serial
    com_ws.Range(DATE_CELLS["paper_table_date"]).Value = reference_serial
    com_ws.Range(DATE_CELLS["market_table_date"]).Value = reference_serial
    com_ws.Range(DATE_CELLS["premium_table_date"]).Value = reference_serial
    _write_report_title_com(com_ws, report_date)

    for item in NUMBER_FIELDS:
        value = parse_optional_number(inputs.get(item.key))
        _write_com_value(com_ws, item.cell, value)
        if item.label_cell and item.label_template:
            rendered = item.label_template.format(
                month=pricing_month.abbr,
                value=_format_paper_label(value),
            )
            _write_com_value(com_ws, item.label_cell, rendered)

    if not skip_shift:
        _write_worldwide_new_point(
            lambda address, value: _write_com_value(com_ws, address, value),
            inputs,
        )

    comment_inputs = dict(inputs)
    worldwide = str(inputs.get(KOREA_WORLDWIDE_COMMENT_KEY, "") or "").strip()
    comment_inputs[KOREA_WORLDWIDE_COMMENT_KEY] = worldwide or inputs.get(KOREA_COMMENT_SOURCE_KEY, "")
    for item in COMMENT_FIELDS:
        lines = compose_comment_lines(comment_inputs.get(item.key, ""), len(item.cells))
        for address, line in zip(item.cells, lines):
            _write_com_value(com_ws, address, line)

    for item in STRATEGY_FIELDS:
        _write_com_value(com_ws, item.cell, compose_strategy_text(item.prefix, inputs.get(item.key, "")))

    _write_com_value(
        com_ws,
        SALES_ASSUMPTION_HEADER_CELL,
        SALES_ASSUMPTION_HEADER_TEMPLATE.format(month=pricing_month.full_name),
    )
    _overlay_extra_cells_com(com_ws, extra_cells)


def _validate_mapping_com(com_ws) -> None:
    mismatches = []
    for address, expected in REQUIRED_LABELS.items():
        actual = com_ws.Range(address).Value
        if actual is None:
            mismatches.append(f"{address} (empty, expected '{_expected_label_ui(expected)}')")
            continue
        if str(actual).strip() != expected:
            mismatches.append(f"{address}='{actual}' (expected '{_expected_label_ui(expected)}')")
    if mismatches:
        raise GenerateError("Sheet structure does not match mapping: " + "; ".join(mismatches))


def _insert_index_for_date(sheets: list[SheetInfo], report_date: dt.date) -> int:
    """1-based COM index among original sheets before the copy is inserted."""
    for sheet in sheets:
        if sheet.date < report_date:
            return sheet.index + 1
    return len(sheets) + 1


def generate_market_report(
    report_date: dt.date,
    inputs: dict[str, Any],
    *,
    export_pdf: bool = True,
    pricing_month: YearMonth | None = None,
    extra_cells: dict[str, Any] | None = None,
    data_reference_date: dt.date | None = None,
) -> GenerateResult:
    try:
        validate_report_date(report_date)
    except WorkingDayError as exc:
        raise GenerateError(str(exc)) from exc

    try:
        template = find_template()
    except FileNotFoundError as exc:
        raise GenerateError(str(exc)) from exc

    data_reference_date = data_reference_date or previous_week_last_working_day(report_date)
    month = pricing_month or default_pricing_month(report_date)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    excel_path = output_excel_path(report_date)
    pdf_path = session_output_dir() / format_pdf_filename(report_date)
    new_sheet_name = format_sheet_name(report_date)

    # Prefer the same-date output workbook when it already has the sheet.
    existing_sheet: str | None = None
    source_for_copy: Path | None = None
    if excel_path.exists():
        try:
            existing_sheet = sheet_exists_for_date(_sheet_names(excel_path), report_date)
            if existing_sheet:
                source_for_copy = excel_path
        except Exception:
            existing_sheet = None

    source = resolve_source_workbook()
    source_names = _sheet_names(source)
    if existing_sheet is None:
        existing_sheet = sheet_exists_for_date(source_names, report_date)
        if existing_sheet:
            source_for_copy = source

    is_update = existing_sheet is not None
    previous = find_previous_sheet(source_names, report_date)
    if not is_update and previous is None:
        # Fallback: look for previous sheet inside the dated output if present.
        if excel_path.exists():
            previous = find_previous_sheet(_sheet_names(excel_path), report_date)
            if previous is not None:
                source = excel_path
                source_names = _sheet_names(source)
    if not is_update and previous is None:
        raise GenerateError("Previous Singapore working-day report sheet was not found.")

    # Never write into templates/. Always operate on the dated output file.
    if excel_path.resolve() == template.resolve():
        raise GenerateError("The original template Excel file cannot be modified.")

    created_fresh_copy = False
    if is_update:
        assert existing_sheet is not None
        assert source_for_copy is not None
        if source_for_copy.resolve() == template.resolve():
            shutil.copy2(source_for_copy, excel_path)
            created_fresh_copy = True
        elif source_for_copy.resolve() != excel_path.resolve():
            shutil.copy2(source_for_copy, excel_path)
            created_fresh_copy = True
        # else: update the existing dated output in place
        previous_sheet_label = existing_sheet
        target_sheet_name = existing_sheet
    else:
        assert previous is not None
        if source.resolve() == template.resolve() or source.resolve() != excel_path.resolve():
            shutil.copy2(source, excel_path)
            created_fresh_copy = True
        previous_sheet_label = previous.name
        target_sheet_name = new_sheet_name

    if excel_path.resolve() == template.resolve():
        raise GenerateError("The original template Excel file cannot be modified.")

    warnings: list[str] = []
    page_count: int | None = None

    try:
        if excel_com_available():
            page_count = _generate_with_com(
                excel_path=excel_path,
                pdf_path=pdf_path if export_pdf else None,
                report_date=report_date,
                data_reference_date=data_reference_date,
                previous_name=None if is_update else previous.name,
                target_sheet_name=target_sheet_name,
                inputs=inputs,
                pricing_month=month,
                is_update=is_update,
                extra_cells=extra_cells,
            )
            if export_pdf and page_count is not None and page_count != EXPECTED_PDF_PAGES:
                warnings.append(
                    f"PDF has {page_count} page(s). The reference layout is {EXPECTED_PDF_PAGES} pages. "
                    "Check Print Area / Page Setup."
                )
        else:
            warnings.append(
                "Excel COM is unavailable, so the sheet was copied with openpyxl. Charts and images may not be fully preserved."
            )
            _generate_with_openpyxl(
                excel_path=excel_path,
                report_date=report_date,
                data_reference_date=data_reference_date,
                previous_name=None if is_update else previous.name,
                target_sheet_name=target_sheet_name,
                inputs=inputs,
                pricing_month=month,
                is_update=is_update,
                extra_cells=extra_cells,
            )
            if export_pdf:
                try:
                    _, page_count = export_report_pdf_cloud(
                        pdf_path=pdf_path,
                        report_date=report_date,
                        data_reference_date=data_reference_date,
                        pricing_month=month.label(),
                        sheet_name=target_sheet_name,
                        inputs=inputs,
                        extra_cells=extra_cells,
                    )
                    warnings.append(
                        "PDF was generated with the cloud renderer because Microsoft Excel COM is unavailable."
                    )
                except Exception as exc:
                    pdf_path = None
                    raise GenerateError(f"Cloud PDF export failed: {exc}") from exc
            else:
                pdf_path = None
    except Exception:
        # On create failure after a fresh copy, remove the incomplete output.
        # Never delete an update target, templates, or a pre-existing workbook.
        if (
            not is_update
            and created_fresh_copy
            and excel_path.exists()
            and excel_path.resolve() != template.resolve()
            and excel_path.resolve() != source.resolve()
        ):
            excel_path.unlink()
        if (
            not is_update
            and created_fresh_copy
            and pdf_path is not None
            and pdf_path.exists()
        ):
            pdf_path.unlink()
        raise

    created_pdf = pdf_path if (export_pdf and pdf_path and pdf_path.exists()) else None
    return GenerateResult(
        excel_path=excel_path,
        pdf_path=created_pdf,
        sheet_name=target_sheet_name,
        source_path=source_for_copy if is_update and source_for_copy else source,
        previous_sheet=previous_sheet_label,
        warnings=warnings,
        pdf_page_count=page_count,
        pricing_month=month.label(),
        is_update=is_update,
    )


def _generate_with_com(
    *,
    excel_path: Path,
    pdf_path: Path | None,
    report_date: dt.date,
    data_reference_date: dt.date,
    previous_name: str | None,
    target_sheet_name: str,
    inputs: dict[str, Any],
    pricing_month: YearMonth,
    is_update: bool,
    extra_cells: dict[str, Any] | None = None,
) -> int | None:
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    page_count: int | None = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        excel.ScreenUpdating = False
        workbook = excel.Workbooks.Open(str(excel_path.resolve()), UpdateLinks=0, ReadOnly=False)

        names = [workbook.Worksheets(i).Name for i in range(1, workbook.Worksheets.Count + 1)]

        if is_update:
            existing = sheet_exists_for_date(names, report_date)
            if existing is None:
                raise GenerateError(f"Could not find the sheet to update: {target_sheet_name}")
            new_ws = workbook.Worksheets(existing)
        else:
            if sheet_exists_for_date(names, report_date):
                raise GenerateError(
                    f"Sheet '{format_sheet_name(report_date)}' already exists and cannot be created again."
                )
            if not previous_name:
                raise GenerateError("Previous Singapore working-day report sheet was not found.")
            source_ws = workbook.Worksheets(previous_name)
            sheets = list_report_sheets(names)
            before_index = _insert_index_for_date(sheets, report_date)
            original_count = workbook.Worksheets.Count
            if before_index <= original_count:
                source_ws.Copy(Before=workbook.Worksheets(before_index))
                new_ws = workbook.Worksheets(before_index)
            else:
                source_ws.Copy(After=workbook.Worksheets(original_count))
                new_ws = workbook.Worksheets(workbook.Worksheets.Count)
            try:
                new_ws.Name = target_sheet_name
            except Exception as exc:
                raise GenerateError(f"Could not rename the sheet: {exc}") from exc

        _validate_mapping_com(new_ws)
        _apply_values_com(
            new_ws, report_date, inputs, pricing_month, data_reference_date, extra_cells
        )

        workbook.Save()
        if pdf_path is not None:
            try:
                _, page_count = prepare_sheet_and_export_pdf(excel, workbook, new_ws, pdf_path)
            except PdfExportError as exc:
                raise GenerateError(str(exc)) from exc
            workbook.Save()
        else:
            excel.Calculate()
            workbook.Save()
        return page_count
    except GenerateError:
        raise
    except Exception as exc:
        action = "update" if is_update else "create"
        raise GenerateError(f"Excel {action} failed: {exc}") from exc
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def _generate_with_openpyxl(
    *,
    excel_path: Path,
    report_date: dt.date,
    data_reference_date: dt.date,
    previous_name: str | None,
    target_sheet_name: str,
    inputs: dict[str, Any],
    pricing_month: YearMonth,
    is_update: bool,
    extra_cells: dict[str, Any] | None = None,
) -> None:
    workbook = load_workbook(excel_path)
    try:
        if is_update:
            existing = sheet_exists_for_date(list(workbook.sheetnames), report_date)
            if existing is None:
                raise GenerateError(f"Could not find the sheet to update: {target_sheet_name}")
            new_ws = workbook[existing]
        else:
            if sheet_exists_for_date(list(workbook.sheetnames), report_date):
                raise GenerateError(
                    f"Sheet '{format_sheet_name(report_date)}' already exists and cannot be created again."
                )
            if not previous_name or previous_name not in workbook.sheetnames:
                raise GenerateError("Previous Singapore working-day report sheet was not found.")

            source_ws = workbook[previous_name]
            new_ws = workbook.copy_worksheet(source_ws)
            new_ws.title = target_sheet_name

            names = list(workbook.sheetnames)
            sheets = list_report_sheets([name for name in names if name != target_sheet_name])
            before_index = _insert_index_for_date(sheets, report_date) - 1
            current_index = workbook.sheetnames.index(target_sheet_name)
            target = min(max(before_index, 0), len(workbook.sheetnames) - 1)
            workbook.move_sheet(new_ws, offset=target - current_index)

        _validate_mapping(new_ws)
        _apply_values(
            new_ws, report_date, inputs, pricing_month, data_reference_date, extra_cells
        )
        workbook.save(excel_path)
    finally:
        workbook.close()
